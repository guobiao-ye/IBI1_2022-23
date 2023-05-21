# Please make sure the terminal installed openpyxl library before running the code!
import xml.dom.minidom as dom
import openpyxl

def find_child_nodes(term, all_terms):
    child_nodes = []
    term_id = term.getElementsByTagName('id')[0].childNodes[0].data
    for t in all_terms:
        is_a = t.getElementsByTagName('is_a')
        for isa in is_a:
            if isa.childNodes[0].data == term_id:
                child_nodes.append(t)
                child_nodes.extend(find_child_nodes(t, all_terms))
    return child_nodes

# Loading XML file
xml_doc = dom.parse('go_obo.xml')
terms = xml_doc.getElementsByTagName('term')

# Create an Excel spreadsheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set header
sheet['A1'] = 'GO id'
sheet['B1'] = 'Term Name'
sheet['C1'] = 'Definition String'
sheet['D1'] = 'Number of Child Nodes'

# Find and extract relevant information
row = 2  # Store data from the second row
for term in terms:
    defstr = term.getElementsByTagName('defstr')[0].childNodes[0].data
    if 'autophagosome' in defstr:
        go_id = term.getElementsByTagName('id')[0].childNodes[0].data
        term_name = term.getElementsByTagName('name')[0].childNodes[0].data
        
        # Find the number of child nodes
        child_nodes = find_child_nodes(term, terms)
        num_child_nodes = len(child_nodes)

        # Store data in an Excel spreadsheet
        sheet.cell(row=row, column=1).value = go_id
        sheet.cell(row=row, column=2).value = term_name
        sheet.cell(row=row, column=3).value = defstr
        sheet.cell(row=row, column=4).value = num_child_nodes

        row += 1

# Save the Excel spreadsheet
workbook.save('autophagosome.xlsx')